"""テキスト層を持つ請求書PDFを読み、検算に通ったものだけをExcelへ追記する。

設計の前提（緩めないこと）:
  * 対象はテキスト埋め込みPDFのみ。スキャン・写真・手書きは検出して拒否する。
  * 1発行元・1レイアウト専用。フォーマット定義は formats/*.json に外出しする。
  * 検算に1つでも失敗した請求書は、Excelに一切書かない。部分的な転記をしない。

検算に通らないものを人間に返すのが、このツールの価値である。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook

# テキスト層ありと判定する最小文字数。これ未満はスキャンPDFの疑いとして拒否する。
MIN_TEXT_CHARS = 50

COLUMNS = [
    "請求書番号",
    "発行日",
    "請求先",
    "品名",
    "数量",
    "単価",
    "金額",
    "小計",
    "消費税",
    "合計金額",
    "元ファイル",
]


class ExtractionError(Exception):
    """転記を中止すべき理由。メッセージはそのまま利用者に見せる。"""


@dataclass
class LineItem:
    name: str
    qty: int
    unit: int
    amount: int


@dataclass
class Invoice:
    invoice_no: str
    issue_date: str
    customer: str
    subtotal: int
    tax: int
    total: int
    items: list[LineItem] = field(default_factory=list)
    source: str = ""
    # 明細らしき見た目なのに定義に当てはまらなかった行。項目が増えた月の検知に使う。
    unparsed: list[str] = field(default_factory=list)


def to_int(s: str) -> int:
    return int(s.replace(",", "").replace("¥", "").strip())


def read_text(pdf_path: Path) -> str:
    """PDFのテキスト層を読む。テキストが薄い場合はスキャン疑いとして拒否する。"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        raise ExtractionError(f"PDFを開けませんでした: {e}") from e

    if len(text.strip()) < MIN_TEXT_CHARS:
        raise ExtractionError(
            "テキスト層が見つかりません。スキャン・写真・手書きの請求書は"
            "このツールの対象外です（誤認識を検算できないため）。"
            "発行元にテキスト形式のPDFを依頼してください。"
        )
    return text


# 金額らしいトークン（桁区切りのある数、または3桁以上の数）。マイナスも拾う。
MONEY = re.compile(r"-?\d{1,3}(?:,\d{3})+|-?\d{3,}")


def looks_like_row(line: str) -> bool:
    """明細欄の中で、金額を含むのに定義に当てはまらなかった行か。"""
    return bool(MONEY.search(line))


def item_region(lines: list[str], region: dict | None) -> list[str]:
    """明細欄だけを切り出す。宛先や合計欄を明細と取り違えないための境界。"""
    if not region:
        return lines
    start, end = region.get("start"), region.get("end")
    a = next((i + 1 for i, l in enumerate(lines) if start and start in l), 0)
    b = next((i for i, l in enumerate(lines) if i >= a and end and end in l), len(lines))
    return lines[a:b]


def find_one(pattern: str, text: str, label: str) -> str:
    m = re.search(pattern, text, re.MULTILINE)
    if not m:
        raise ExtractionError(
            f"{label} を読み取れませんでした。"
            "フォーマット定義とPDFのレイアウトが一致していない可能性があります。"
        )
    return m.group(1).strip()


def parse(text: str, fmt: dict, source: str) -> Invoice:
    header = fmt["header"]
    totals = fmt["totals"]

    items: list[LineItem] = []
    unparsed: list[str] = []
    pattern = re.compile(fmt["line_item"])
    excludes = fmt.get("line_item_exclude", [])
    region = item_region(text.splitlines(), fmt.get("line_item_region"))
    for line in region:
        line = line.strip()
        if not line or any(x in line for x in excludes):
            continue
        m = pattern.match(line)
        if m:
            items.append(
                LineItem(
                    name=m.group("name").strip(),
                    qty=to_int(m.group("qty")),
                    unit=to_int(m.group("unit")),
                    amount=to_int(m.group("amount")),
                )
            )
        elif looks_like_row(line):
            unparsed.append(line)

    if not items:
        raise ExtractionError("明細行を1件も読み取れませんでした。")

    return Invoice(
        invoice_no=find_one(header["invoice_no"], text, "請求書番号"),
        issue_date=find_one(header["issue_date"], text, "発行日"),
        customer=find_one(header["customer"], text, "請求先"),
        subtotal=to_int(find_one(totals["subtotal"], text, "小計")),
        tax=to_int(find_one(totals["tax"], text, "消費税")),
        total=to_int(find_one(totals["total"], text, "合計金額")),
        items=items,
        source=source,
        unparsed=unparsed,
    )


def verify(inv: Invoice) -> list[str]:
    """検算。返り値が空でなければ、その請求書は転記しない。"""
    problems: list[str] = []

    for item in inv.items:
        expected = item.qty * item.unit
        if expected != item.amount:
            problems.append(
                f"明細「{item.name}」: 数量×単価={expected:,} だが 金額={item.amount:,}"
                f"（差額 {item.amount - expected:+,}）"
            )

    items_sum = sum(i.amount for i in inv.items)
    if items_sum != inv.subtotal:
        problems.append(
            f"明細合計={items_sum:,} だが 小計={inv.subtotal:,}"
            f"（差額 {items_sum - inv.subtotal:+,}）"
        )
        if inv.unparsed:
            problems.append(
                "読み取れなかった行があります（請求書に新しい項目が増えた可能性）: "
                + " / ".join(inv.unparsed)
            )

    if inv.subtotal + inv.tax != inv.total:
        problems.append(
            f"小計+消費税={inv.subtotal + inv.tax:,} だが 合計金額={inv.total:,}"
            f"（差額 {inv.subtotal + inv.tax - inv.total:+,}）"
        )

    return problems


def rows_of(inv: Invoice) -> list[list]:
    """1請求書ぶんの行。先頭行にだけ合計欄を入れる（Excel/TSVで共通）。"""
    out = []
    for i, item in enumerate(inv.items):
        first = i == 0
        out.append([
            inv.invoice_no, inv.issue_date, inv.customer,
            item.name, item.qty, item.unit, item.amount,
            inv.subtotal if first else None,
            inv.tax if first else None,
            inv.total if first else None,
            inv.source if first else None,
        ])
    return out


def write_tsv(path: Path, invoices: list[Invoice]) -> None:
    """スプレッドシートへ直接貼り付けるためのTSV。Excelを開かずに使える。"""
    lines = ["\t".join(COLUMNS)]
    for inv in invoices:
        for row in rows_of(inv):
            lines.append("\t".join("" if v is None else str(v) for v in row))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def load_sheet(excel: Path, sheet: str):
    if excel.exists():
        wb = load_workbook(excel)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        if ws.max_row == 1 and all(c.value is None for c in ws[1]):
            ws.append(COLUMNS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(COLUMNS)
    return wb, ws


def existing_numbers(ws) -> set[str]:
    """すでに転記済みの請求書番号。二重計上を防ぐ。"""
    return {row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]}


def append(ws, inv: Invoice) -> None:
    for row in rows_of(inv):
        ws.append(row)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("pdfs", nargs="+", help="請求書PDF（複数可）")
    ap.add_argument("--format", required=True, help="フォーマット定義JSON")
    ap.add_argument("--excel", help="追記先のExcelファイル")
    ap.add_argument("--tsv", help="スプレッドシートへ貼り付ける用のTSVを書き出す")
    ap.add_argument("--sheet", default="請求明細", help="シート名")
    ap.add_argument("--dry-run", action="store_true", help="Excelに書かず検算結果だけ表示")
    args = ap.parse_args()

    if not args.excel and not args.tsv:
        ap.error("--excel か --tsv のどちらかを指定してください。")

    fmt = json.loads(Path(args.format).read_text(encoding="utf-8"))
    excel = Path(args.excel) if args.excel else None
    if excel:
        wb, ws = load_sheet(excel, args.sheet)
        already = existing_numbers(ws)
    else:
        wb = ws = None
        already = set()

    ok: list[Invoice] = []
    skipped: list[tuple[str, str]] = []

    for raw in args.pdfs:
        path = Path(raw)
        try:
            inv = parse(read_text(path), fmt, path.name)
        except ExtractionError as e:
            skipped.append((path.name, str(e)))
            continue

        if inv.invoice_no in already:
            skipped.append((path.name, f"請求書番号 {inv.invoice_no} は転記済みです（二重計上を防止）"))
            continue

        problems = verify(inv)
        if problems:
            skipped.append((path.name, "検算不一致 — " + " / ".join(problems)))
            continue

        ok.append(inv)
        already.add(inv.invoice_no)

    for inv in ok:
        print(f"OK   {inv.source}: {inv.invoice_no} 合計 ¥{inv.total:,}（明細{len(inv.items)}行）")
    for name, reason in skipped:
        print(f"要確認 {name}: {reason}")

    if args.dry_run:
        print("\n[dry-run] Excelには書き込んでいません。")
        return 1 if skipped else 0

    if ok:
        if excel:
            for inv in ok:
                append(ws, inv)
            wb.save(excel)
            print(f"\n{len(ok)}件を {excel} に追記しました。")
        if args.tsv:
            write_tsv(Path(args.tsv), ok)
            print(f"{len(ok)}件を {args.tsv} に書き出しました（スプレッドシートに貼り付けられます）。")
    else:
        print("\n転記できた請求書はありません。出力は変更していません。")

    if skipped:
        print(f"{len(skipped)}件は転記していません。上の理由を確認してください。")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
